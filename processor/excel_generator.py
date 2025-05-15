"""
Functions for generating Excel report output.
"""

import pandas as pd
import math
from datetime import datetime
from .connection_processing import get_midspan_proposed_heights
from .utils import calculate_bearing
from .height_utils import get_pole_primary_neutral_heights, get_attacher_ground_clearance # Added
# format_height_feet_inches is also in height_utils but not directly used here, it's used by the other two.

def create_output_excel(output_excel_path, df, job_data):
    """
    Create a well-formatted Excel report from the processed data with enhanced formatting.
    Follows the format with multiple rows per pole (one for each attacher) and organized by pole pairs.
    
    Args:
        output_excel_path (str): Path where the Excel file will be saved
        df (pd.DataFrame): The processed data to include in the report
        job_data (dict): The original Katapult JSON data
        
    Returns:
        None
    """
    if df.empty:
        print("No data available to create Excel report")
        return
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
        from openpyxl.utils import get_column_letter
        
        # Create Excel writer with openpyxl engine
        with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
            # Access the workbook
            workbook = writer.book
            
            # Create a new sheet for the report
            workbook.create_sheet("Make Ready Report")
            main_sheet = workbook["Make Ready Report"]
            
            # ----- Define Excel Styles -----
            header_font = Font(name='Arial', size=11, bold=True, color='000000')
            header_font_white = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            
            # Custom color fills as per requirements
            section_header_fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')  # Light blue
            subheader_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')  # Lighter blue
            underground_fill = PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid')  # Light red
            backspan_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # Blue-gray
            reference_fill = PatternFill(start_color='E2EFD9', end_color='E2EFD9', fill_type='solid')  # Light green
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # ----- Create Multi-Level Header -----
            # Create the top-level header row (merged cells for categories)
            
            # Column A - Connection ID
            main_sheet.cell(row=1, column=1).value = "Connection ID"
            main_sheet.merge_cells('A1:A2')
            
            # Column B - Operation Number
            main_sheet.cell(row=1, column=2).value = "Operation Number"
            main_sheet.merge_cells('B1:B2')
            
            # Column C - Attachment Action
            main_sheet.cell(row=1, column=3).value = "Attachment Action:"
            main_sheet.cell(row=2, column=3).value = "(I)nstalling\n(R)emoving\n(E)xisting"
            
            # Column D - Pole Owner
            main_sheet.cell(row=1, column=4).value = "Pole Owner"
            main_sheet.merge_cells('D1:D2')
            
            # Column E - Pole #
            main_sheet.cell(row=1, column=5).value = "Pole #"
            main_sheet.merge_cells('E1:E2')
            
            # Column F - SCID
            main_sheet.cell(row=1, column=6).value = "SCID"
            main_sheet.merge_cells('F1:F2')
            
            # Column G - Pole Structure
            main_sheet.cell(row=1, column=7).value = "Pole Structure"
            main_sheet.merge_cells('G1:G2')
            
            # Column H - Proposed Riser
            main_sheet.cell(row=1, column=8).value = "Proposed Riser (Yes/No)"
            main_sheet.merge_cells('H1:H2')
            
            # Column I - Proposed Guy
            main_sheet.cell(row=1, column=9).value = "Proposed Guy (Yes/No)"
            main_sheet.merge_cells('I1:I2')
            
            # Column J - PLA Percentage
            main_sheet.cell(row=1, column=10).value = "PLA (%) with proposed attachment"
            main_sheet.merge_cells('J1:J2')
            
            # Column K - Construction Grade
            main_sheet.cell(row=1, column=11).value = "Construction Grade of Analysis"
            main_sheet.merge_cells('K1:K2')
            
            # Column L-M - Existing Mid-Span Data
            main_sheet.cell(row=1, column=12).value = "Existing Mid-Span Data"
            main_sheet.merge_cells('L1:M1')
            main_sheet.cell(row=2, column=12).value = "Height Lowest Com"
            main_sheet.cell(row=2, column=13).value = "Height Lowest CPS Electrical"
            
            # Column N-T - Make Ready Data (Expanded)
            main_sheet.cell(row=1, column=14).value = "Make Ready Data"
            main_sheet.merge_cells('N1:T1') # Changed from N1:Q1
            main_sheet.cell(row=2, column=14).value = "Attacher Name" # Changed from "Attacher Description"
            main_sheet.cell(row=2, column=15).value = "Existing Height"
            main_sheet.cell(row=2, column=16).value = "Proposed Height"
            main_sheet.cell(row=2, column=17).value = "Mid-Span Proposed"
            main_sheet.cell(row=2, column=18).value = "Ground Clearance" # New position
            main_sheet.cell(row=2, column=19).value = "Neutral Height"   # New position
            main_sheet.cell(row=2, column=20).value = "Primary Height"   # New position
            
            # Columns U-X - Movement Information (No change in columns, but position relative to previous headers changes)
            main_sheet.cell(row=1, column=21).value = "Movement Information"
            main_sheet.merge_cells('U1:X1') # Stays U-X
            main_sheet.cell(row=2, column=21).value = "Move Distance"
            main_sheet.cell(row=2, column=22).value = "Direction"
            main_sheet.cell(row=2, column=23).value = "Span Sag"
            main_sheet.cell(row=2, column=24).value = "Notes"
            
            # Apply header styles - section headers (Row 1)
            for row in [1]:
                for col in range(1, 25):  # Apply to all 24 columns
                    cell = main_sheet.cell(row=row, column=col)
                    cell.font = header_font
                    cell.fill = section_header_fill
                    cell.border = thin_border
                    cell.alignment = centered_alignment
            
            # Apply subheader styles (Row 2)
            for row in [2]:
                for col in range(1, 25):  # Apply to all 24 columns
                    cell = main_sheet.cell(row=row, column=col)
                    cell.font = header_font
                    cell.fill = subheader_fill
                    cell.border = thin_border
                    cell.alignment = centered_alignment
            
            # Process data by node/connection pairs
            current_row = 3  # Start after the header rows
            
            # Group by node_id_1 to handle each pole separately
            grouped_by_node = df.groupby('node_id_1')
            
            for node_id, node_group in grouped_by_node:
                # For each pole, process its first connection only (if we've already processed this pole, skip it)
                if node_group.empty:
                    continue
                
                # Get the first record for this node
                first_record = node_group.iloc[0]
                current_node_id = first_record.get('node_id_1') # Get current node_id for height lookups

                # Get pole-specific primary and neutral heights once per pole
                pole_specific_heights = {}
                if current_node_id and job_data:
                    pole_specific_heights = get_pole_primary_neutral_heights(current_node_id, job_data)
                
                pole_primary_h_str = pole_specific_heights.get('primary_height', '')
                pole_neutral_h_str = pole_specific_heights.get('neutral_height', '')
                
                # Write basic pole information
                operation_number = first_record.get('operation_number')
                attachment_action = first_record.get('attachment_action', '(E)xisting')
                pole_owner = first_record.get('pole_owner', '')
                pole_number = first_record.get('pole_tag_1', '')
                pole_structure = first_record.get('pole_structure', '')
                proposed_riser = first_record.get('proposed_riser', 'NO')
                proposed_guy = first_record.get('proposed_guy', 'NO')
                pla_percentage = first_record.get('pla_percentage', '')
                construction_grade = first_record.get('construction_grade', '')
                lowest_com = first_record.get('lowest_com_height', '')
                lowest_cps = first_record.get('lowest_cps_height', '')
                pole_tag_1 = first_record.get('pole_tag_1', '')
                pole_tag_2 = first_record.get('pole_tag_2', '')
                
                # Insert "From Pole" header row
                from_pole_row = current_row
                main_sheet.cell(row=from_pole_row, column=10).value = "From Pole"
                main_sheet.merge_cells(f'J{from_pole_row}:K{from_pole_row}')
                main_sheet.cell(row=from_pole_row, column=10).font = header_font
                main_sheet.cell(row=from_pole_row, column=10).fill = section_header_fill
                main_sheet.cell(row=from_pole_row, column=10).alignment = centered_alignment
                main_sheet.cell(row=from_pole_row, column=10).border = thin_border
                main_sheet.cell(row=from_pole_row, column=11).fill = section_header_fill
                main_sheet.cell(row=from_pole_row, column=11).border = thin_border
                
                # Add pole tag info in the next row
                tag_row = from_pole_row + 1
                main_sheet.cell(row=tag_row, column=10).value = pole_tag_1
                main_sheet.cell(row=tag_row, column=11).value = pole_tag_2
                
                # If we have a valid connection, insert the reference direction
                if pole_tag_2:
                    # Get the connection data
                    connection_id = first_record.get('connection_id')
                    reference_row = from_pole_row + 2
                    
                    # Insert the reference direction row with proper reference formatting (light green)
                    main_sheet.cell(row=reference_row, column=12).value = f"Reference or Other_pole [cardinal direction] to {pole_tag_2}"
                    main_sheet.merge_cells(f'L{reference_row}:O{reference_row}')
                    main_sheet.cell(row=reference_row, column=12).font = header_font
                    main_sheet.cell(row=reference_row, column=12).fill = reference_fill
                    main_sheet.cell(row=reference_row, column=12).alignment = centered_alignment
                    main_sheet.cell(row=reference_row, column=12).border = thin_border

                    # Apply reference formatting to the whole row
                    for col in range(1, 16):
                        if col not in range(12, 16):  # Skip the already merged cells
                            cell = main_sheet.cell(row=reference_row, column=col)
                            cell.fill = reference_fill
                            cell.border = thin_border
                    
                    # Start processing attachers after the reference row
                    attacher_start_row = reference_row + 1
                else:
                    # If no connection, start processing attachers right after the tag row
                    attacher_start_row = tag_row + 1
                
                # Get the attacher data for this pole
                attachers_data = first_record.get('attachers_data', {})
                main_attachers = attachers_data.get('main_attachers', [])
                
                # If there are no attachers, add at least one row with the basic pole data
                if not main_attachers:
                    # Write basic pole data to all 24 columns
                    row = attacher_start_row
                    connection_id = first_record.get('connection_id', '')
                    scid = first_record.get('scid_1', '')

                    # First section - identification columns
                    main_sheet.cell(row=row, column=1).value = connection_id
                    main_sheet.cell(row=row, column=2).value = operation_number
                    main_sheet.cell(row=row, column=3).value = attachment_action
                    main_sheet.cell(row=row, column=4).value = pole_owner
                    main_sheet.cell(row=row, column=5).value = pole_number
                    main_sheet.cell(row=row, column=6).value = scid
                    main_sheet.cell(row=row, column=7).value = pole_structure
                    main_sheet.cell(row=row, column=8).value = proposed_riser
                    main_sheet.cell(row=row, column=9).value = proposed_guy
                    main_sheet.cell(row=row, column=10).value = pla_percentage
                    main_sheet.cell(row=row, column=11).value = construction_grade
                    main_sheet.cell(row=row, column=12).value = lowest_com
                    main_sheet.cell(row=row, column=13).value = lowest_cps
                    
                    # Apply borders and styling to all cells in the row
                    for col in range(1, 25):
                        cell = main_sheet.cell(row=row, column=col)
                        cell.border = thin_border
                        cell.alignment = centered_alignment  # Center all cells
                    
                    # Move to next row
                    current_row = row + 1
                else:
                    # Add multiple rows - one for each attacher
                    for idx, attacher in enumerate(main_attachers):
                        row = attacher_start_row + idx
                        attacher_name = attacher.get('name', '')
                        existing_height = attacher.get('existing_height', '')
                        proposed_height = attacher.get('proposed_height', '')
                        
                        # Get midspan height for this attacher
                        midspan_height = ""
                        if first_record.get('connection_id') and job_data: # ensure job_data is available
                            midspan_height = get_midspan_proposed_heights(job_data, first_record['connection_id'], attacher_name)
                        
                        # Get attacher specific ground clearance
                        attacher_gc_str = ""
                        if current_node_id and attacher_name and job_data:
                             attacher_gc_str = get_attacher_ground_clearance(current_node_id, attacher_name, job_data)

                        # Check for special row types - underground, backspan, or reference
                        connection_type = first_record.get('connection_type', '').lower()
                        is_underground = 'underground' in connection_type
                        is_backspan = attacher.get('is_backspan', False)
                        is_reference = attacher.get('is_reference', False)
                        
                        # Apply special formatting based on row type
                        row_fill = None
                        if is_underground:
                            row_fill = underground_fill  # Light red
                        elif is_backspan:
                            row_fill = backspan_fill     # Blue-gray
                        elif is_reference:
                            row_fill = reference_fill    # Light green
                        
                        # Prepare data for movement columns
                        move_distance = ""
                        move_direction = ""
                        if existing_height and proposed_height and existing_height != proposed_height:
                            try:
                                # Parse heights from feet-inches format (e.g., "45'-6"")
                                existing_parts = existing_height.replace('"', '').split("'")
                                proposed_parts = proposed_height.replace('"', '').split("'")
                                existing_inches = (int(existing_parts[0]) * 12) + (int(existing_parts[1]) if len(existing_parts) > 1 else 0)
                                proposed_inches = (int(proposed_parts[0]) * 12) + (int(proposed_parts[1]) if len(proposed_parts) > 1 else 0)
                                
                                # Calculate movement information
                                move_distance = f"{abs(proposed_inches - existing_inches)}\""
                                move_direction = "Up" if proposed_inches > existing_inches else "Down"
                            except:
                                pass  # Ignore if parsing fails
                        
                        # Write basic pole data (only for the first attacher row)
                        if idx == 0:
                            connection_id = first_record.get('connection_id', '')
                            scid = first_record.get('scid_1', '')
                            
                            main_sheet.cell(row=row, column=1).value = connection_id
                            main_sheet.cell(row=row, column=2).value = operation_number
                            main_sheet.cell(row=row, column=3).value = attachment_action
                            main_sheet.cell(row=row, column=4).value = pole_owner
                            main_sheet.cell(row=row, column=5).value = pole_number
                            main_sheet.cell(row=row, column=6).value = scid
                            main_sheet.cell(row=row, column=7).value = pole_structure
                            main_sheet.cell(row=row, column=8).value = proposed_riser
                            main_sheet.cell(row=row, column=9).value = proposed_guy
                            main_sheet.cell(row=row, column=10).value = pla_percentage
                            main_sheet.cell(row=row, column=11).value = construction_grade
                            main_sheet.cell(row=row, column=12).value = lowest_com
                            main_sheet.cell(row=row, column=13).value = lowest_cps
                        
                        # Write attacher data
                        main_sheet.cell(row=row, column=14).value = attacher_name # Attacher Name
                        main_sheet.cell(row=row, column=15).value = existing_height # Existing Height
                        main_sheet.cell(row=row, column=16).value = proposed_height # Proposed Height
                        main_sheet.cell(row=row, column=17).value = midspan_height # Mid-Span Proposed
                        main_sheet.cell(row=row, column=18).value = attacher_gc_str # Ground Clearance
                        main_sheet.cell(row=row, column=19).value = pole_neutral_h_str # Neutral Height (for the pole)
                        main_sheet.cell(row=row, column=20).value = pole_primary_h_str # Primary Height (for the pole)
                        
                        # Write movement information
                        main_sheet.cell(row=row, column=21).value = move_distance
                        main_sheet.cell(row=row, column=22).value = move_direction
                        
                        # Apply borders, centering, and special fill to all cells in the row
                        for col in range(1, 25):
                            cell = main_sheet.cell(row=row, column=col)
                            cell.border = thin_border
                            cell.alignment = centered_alignment  # Center all cells as required
                            
                            # Apply special background color if this is a special row type
                            if row_fill:
                                cell.fill = row_fill
                    
                    # Update current row for next pole
                    current_row = attacher_start_row + len(main_attachers)
                
                # Add a 'To Pole' section if this is the end of processing for this pole
                if pole_tag_2:
                    to_pole_row = current_row
                    main_sheet.cell(row=to_pole_row, column=10).value = "To Pole"
                    main_sheet.merge_cells(f'J{to_pole_row}:K{to_pole_row}')
                    main_sheet.cell(row=to_pole_row, column=10).font = header_font
                    main_sheet.cell(row=to_pole_row, column=10).fill = section_header_fill
                    main_sheet.cell(row=to_pole_row, column=10).alignment = centered_alignment
                    main_sheet.cell(row=to_pole_row, column=10).border = thin_border
                    main_sheet.cell(row=to_pole_row, column=11).fill = section_header_fill
                    main_sheet.cell(row=to_pole_row, column=11).border = thin_border
                    
                    # Add pole tags for To Pole
                    to_tag_row = to_pole_row + 1
                    main_sheet.cell(row=to_tag_row, column=10).value = pole_tag_1
                    main_sheet.cell(row=to_tag_row, column=11).value = pole_tag_2
                    
                    # Update current row
                    current_row = to_tag_row + 1
                
                # Add a blank row after each pole's data for better readability
                current_row += 1
            
            # Set column widths for all 24 columns
            column_widths = {
                'A': 15,  # Connection ID
                'B': 15,  # Operation Number
                'C': 15,  # Attachment Action
                'D': 12,  # Pole Owner
                'E': 12,  # Pole #
                'F': 12,  # SCID
                'G': 15,  # Pole Structure
                'H': 18,  # Proposed Riser
                'I': 18,  # Proposed Guy
                'J': 22,  # PLA Percentage
                'K': 22,  # Construction Grade
                'L': 18,  # Height Lowest Com
                'M': 22,  # Height Lowest CPS Electrical
                'N': 25,  # Attacher Description
                'O': 15,  # Existing Height
                'P': 15,  # Proposed Height
                'Q': 15,  # Mid-Span Proposed
                'R': 15,  # Ground Clearance
                'S': 15,  # Neutral Height
                'T': 15,  # Primary Height
                'U': 12,  # Move Distance
                'V': 10,  # Direction
                'W': 12,  # Span Sag
                'X': 20,  # Notes
            }
            
            for col, width in column_widths.items():
                main_sheet.column_dimensions[col].width = width
            
            # Freeze the header rows
            main_sheet.freeze_panes = 'A3'
            
            # Apply alternating row colors to all 24 columns
            for row in range(3, current_row):
                if row % 2 == 0:
                    for col in range(1, 25):  # Apply to all 24 columns
                        cell = main_sheet.cell(row=row, column=col)
                        if cell.fill.start_color.index == '00000000':  # Only apply if no fill already set
                            cell.fill = PatternFill(start_color='E9EDF1', end_color='E9EDF1', fill_type='solid')
            
            # ----- Create Summary Sheet -----
            
            # Get job information
            job_name = job_data.get("job_name", "Unknown Job")
            creation_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Count statistics
            pole_count = len(df['node_id_1'].dropna().unique()) if 'node_id_1' in df.columns else 0
            connection_count = len(df)
            
            # Count proposed items
            proposed_count = sum(1 for action in df['attachment_action'] if action == "(I)nstalling") if 'attachment_action' in df.columns else 0
            proposed_riser_count = sum(1 for riser in df['proposed_riser'] if riser.startswith("YES")) if 'proposed_riser' in df.columns else 0
            proposed_guy_count = sum(1 for guy in df['proposed_guy'] if guy.startswith("YES")) if 'proposed_guy' in df.columns else 0
            
            # Create summary data
            summary_data = [
                ["Make Ready Report Summary", ""],
                ["", ""],
                ["Job Information", ""],
                ["Job Name", job_name],
                ["Report Created", creation_date],
                ["", ""],
                ["Statistics", ""],
                ["Total Poles", str(pole_count)],
                ["Total Connections", str(connection_count)],
                ["Poles with Proposed Attachments", str(proposed_count)],
                ["Poles with Proposed Risers", str(proposed_riser_count)],
                ["Poles with Proposed Guys", str(proposed_guy_count)],
                ["", ""],
                ["Notes", ""],
                ["1. This report was generated from Katapult JSON data", ""],
                ["2. Format matches the user-specified Excel format with rows per attacher", ""],
            ]
            
            # Create a new sheet for the summary
            summary_sheet = workbook.create_sheet("Summary", 0)  # Make it the first sheet
            
            # Write summary data
            for row_num, row_data in enumerate(summary_data, 1):
                for col_num, cell_value in enumerate(row_data, 1):
                    summary_sheet.cell(row=row_num, column=col_num).value = cell_value
            
            # Format summary sheet
            # Set column width
            summary_sheet.column_dimensions['A'].width = 30
            summary_sheet.column_dimensions['B'].width = 50
            
            # Format title
            title_cell = summary_sheet.cell(row=1, column=1)
            title_cell.font = Font(name='Arial', size=16, bold=True)
            
            # Format headers with the same section header color as the main sheet
            header_rows = [3, 7, 14]  # Rows with section headers
            for row_num in header_rows:
                cell = summary_sheet.cell(row=row_num, column=1)
                cell.font = Font(name='Arial', size=12, bold=True)
                cell.fill = section_header_fill  # Use the same light blue as main sheet
                
                # Also merge the header cells across the two columns for visual consistency
                summary_sheet.merge_cells(f'A{row_num}:B{row_num}')
                cell.alignment = centered_alignment
            
            # Format data rows
            data_row_ranges = [(4, 5), (8, 12), (15, 16)]  # Ranges of data rows
            for start_row, end_row in data_row_ranges:
                for row_num in range(start_row, end_row + 1):
                    # Make the label cell bold
                    summary_sheet.cell(row=row_num, column=1).font = Font(name='Arial', size=11, bold=True)
                    
                    # Style the value cell
                    value_cell = summary_sheet.cell(row=row_num, column=2)
                    value_cell.font = Font(name='Arial', size=11)
                    
                    # Add alternating row colors
                    if row_num % 2 == 0:
                        for col in range(1, 3):
                            summary_sheet.cell(row=row_num, column=col).fill = PatternFill(
                                start_color='E9EDF1', end_color='E9EDF1', fill_type='solid')
            
            # Make "Summary" the active sheet when opening
            workbook.active = 0
            
        print(f"Excel report successfully created: {output_excel_path}")
        return
        
    except Exception as e:
        print(f"Error creating Excel report: {str(e)}")
        # If error occurs, create a basic report without formatting
        try:
            # Create a simpler format that follows the layout but with minimal formatting
            pole_data = []
            attacher_data = []
            
            for _, row in df.iterrows():
                # Get basic pole data
                pole_info = {
                    'operation_number': row.get('operation_number'),
                    'attachment_action': row.get('attachment_action', '(E)xisting'),
                    'pole_owner': row.get('pole_owner', ''),
                    'pole_number': row.get('pole_tag_1', ''),
                    'pole_structure': row.get('pole_structure', ''),
                    'proposed_riser': row.get('proposed_riser', 'NO'),
                    'proposed_guy': row.get('proposed_guy', 'NO'),
                    'pla_percentage': row.get('pla_percentage', ''),
                    'construction_grade': row.get('construction_grade', ''),
                    'lowest_com_height': row.get('lowest_com_height', ''),
                    'lowest_cps_height': row.get('lowest_cps_height', ''),
                    'pole_tag_1': row.get('pole_tag_1', ''),
                    'pole_tag_2': row.get('pole_tag_2', ''),
                    'connection_id': row.get('connection_id')
                }
                
                pole_data.append(pole_info)
                
                # Get attacher data for this pole
                attachers_data = row.get('attachers_data', {})
                main_attachers = attachers_data.get('main_attachers', [])
                
                if main_attachers:
                    for attacher in main_attachers:
                        attacher_name = attacher.get('name', '')
                        existing_height = attacher.get('existing_height', '')
                        proposed_height = attacher.get('proposed_height', '')
                        
                        # Get midspan height for this attacher
                        midspan_height = ""
                        if row.get('connection_id'):
                            midspan_height = get_midspan_proposed_heights(job_data, row['connection_id'], attacher_name)
                        
                        attacher_info = pole_info.copy()
                        attacher_info['Attacher Description'] = attacher_name
                        attacher_info['Existing'] = existing_height
                        attacher_info['Proposed'] = proposed_height
                        attacher_info['Mid-Span Proposed'] = midspan_height
                        
                        attacher_data.append(attacher_info)
            
            # Create a new dataframe with the expanded attacher data
            if attacher_data:
                output_df = pd.DataFrame(attacher_data)
            else:
                output_df = pd.DataFrame(pole_data)
            
            # Add attacher columns if they don't exist
            for col in ['Attacher Description', 'Existing', 'Proposed', 'Mid-Span Proposed']:
                if col not in output_df.columns:
                    output_df[col] = ""
            
            # Select columns in desired order
            col_order = [
                'operation_number', 'attachment_action', 'pole_owner', 'pole_number',
                'pole_structure', 'proposed_riser', 'proposed_guy', 'pla_percentage',
                'construction_grade', 'lowest_com_height', 'lowest_cps_height',
                'Attacher Description', 'Existing', 'Proposed', 'Mid-Span Proposed'
            ]
            
            # Ensure all columns exist
            for col in col_order:
                if col not in output_df.columns:
                    output_df[col] = ""
            
            # Write to Excel
            with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
                output_df[col_order].to_excel(writer, sheet_name='Make Ready Report', index=False)
                
                # Create a simple summary sheet
                summary_data = {
                    "Item": ["Job Name", "Total Poles", "Total Connections"],
                    "Value": [
                        job_data.get("job_name", "Unknown Job"),
                        len(df['node_id_1'].dropna().unique()) if 'node_id_1' in df.columns else 0,
                        len(df)
                    ]
                }
                pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
            
            print(f"Basic Excel report created due to formatting error: {output_excel_path}")
        except Exception as backup_error:
            print(f"Failed to create even basic Excel report: {str(backup_error)}")
